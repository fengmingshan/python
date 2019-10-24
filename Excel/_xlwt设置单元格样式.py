# -*- coding: utf-8 -*-
# @Author: Administrator
# @Date:   2019-10-24 14:44:41
# @Last Modified by:   Administrator
# @Last Modified time: 2019-10-24 17:42:55

import pandas as pd
import os
import numpy as np
from openpyxl.styles import Alignment,Font
from openpyxl import Workbook, load_workbook

alignment=Alignment(horizontal='general',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=True,
                    indent=0)

font = Font(name='Calibri',
            size=11,
            color='FF000000',
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False)

data_path = 'D:/Test/修改Excel表格样式'
os.chdir(data_path)
print('当前工作目录设置为{path}'.format(path = data_path))

wb = load_workbook('周报.xlsx')

ws = wb.get_sheet_by_name('全部基站详单')

col = ws.column_dimensions['A']

ws['A1'].font = font
ws['A1'].alignment = alignment
ws['A2'].alignment = alignment

wb.save("周报.xlsx")
